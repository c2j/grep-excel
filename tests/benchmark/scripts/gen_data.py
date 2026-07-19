#!/usr/bin/env python3
"""生成基准测试数据：bench.xlsx (200k行) 与 bench.csv（同数据）。
用法: BENCH_DIR=/path python3 gen_data.py  (默认 /tmp/bench)"""
import csv, os, random, datetime
from openpyxl import Workbook

BASE = os.environ.get("BENCH_DIR", "/tmp/bench")
os.makedirs(BASE, exist_ok=True)

random.seed(42)
N = 200_000
DEPTS = ["Engineering", "Sales", "Finance", "HR", "Support", "Marketing", "Legal", "Ops", "Research", "QA"]
CITIES = ["上海", "北京", "深圳", "杭州", "成都", "南京", "武汉", "西安", "苏州", "重庆"]
SURNAMES = ["张", "李", "王", "刘", "陈", "杨", "赵", "黄", "周", "吴"]
GIVEN = ["伟", "芳", "娜", "磊", "静", "强", "洋", "艳", "勇", "军"]
STATUSES = ["正常", "正常", "正常", "正常", "待复核", "ERROR-超时", "ERROR-校验失败"]
NOTES = ["例行巡检通过", "批量导入", "人工录入", "接口同步", "历史迁移数据", "季度结算生成"]
NEEDLE = "ZXQ-7734-异常交易"

header = ["ID", "Name", "Dept", "City", "Level", "Amount", "Date", "Status", "Note", "TraceCode"]
base = datetime.date(2023, 1, 1)

def row(i):
    needle = (i % 701 == 0)  # ~285 行命中
    note = (NEEDLE + "，需人工复核") if needle else random.choice(NOTES)
    return [
        i,
        random.choice(SURNAMES) + random.choice(GIVEN) + str(random.randint(1, 99)),
        random.choice(DEPTS),
        random.choice(CITIES),
        random.randint(1, 15),
        round(random.uniform(100, 99999), 2),
        (base + datetime.timedelta(days=random.randint(0, 900))).strftime("%Y%m%d"),
        random.choice(STATUSES),
        note,
        f"T{random.randint(100000, 999999)}",
    ]

with open(f"{BASE}/bench.csv", "w", newline="", encoding="utf-8") as f:
    w = csv.writer(f)
    w.writerow(header)
    for i in range(1, N + 1):
        w.writerow(row(i))
print("csv done")

random.seed(42)  # 重置以保证与 CSV 完全一致
wb = Workbook(write_only=True)
ws = wb.create_sheet("Sheet1")
ws.append(header)
for i in range(1, N + 1):
    ws.append(row(i))
wb.save(f"{BASE}/bench.xlsx")
print("xlsx done")

# 第二个小文件，用于跨文件 JOIN 演示
wb2 = Workbook(write_only=True)
ws2 = wb2.create_sheet("Sheet1")
ws2.append(["Dept", "DeptName", "Manager"])
for d in DEPTS:
    ws2.append([d, d + "部", random.choice(SURNAMES) + random.choice(GIVEN)])
wb2.save(f"{BASE}/depts.xlsx")
print("depts done")
